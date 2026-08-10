# -*- coding: utf-8 -*-
"""
구성표 엑셀(7시트) + _소재 섹션 폴더 생성.

  python3 make_planning_sheet.py config.json sections.json

sections.json 예시 — 참고링크를 분해해 만든 섹션·구좌 정의
{
  "sections": [
    {"id":"S01","name":"상단 KV","folder":"01_KV","shot":"S01_상단KV.jpg",
     "slots":[
       {"gid":"A01","gname":"KV 배경","item":"배경 이미지","type":"이미지",
        "value":"","file":"kv_bg.jpg","spec":"1920 x 635 (JPG)","guide":"..."},
       {"gid":"A02","gname":"메인 타이틀","item":"1행","type":"텍스트","value":""}
     ]},
    {"id":"S02","name":"앵커 네비게이션","folder":"—","nav":true}
  ]
}

⚠ 이 스크립트는 **새 폴더에서 최초 1회만** 쓴다.
   이미 사용자가 채운 구성표가 있으면 절대 실행하지 않는다 (덮어쓰기 사고).
   기존 파일을 고칠 때는 read_workbook.py 로 읽어 바뀌는 칸만 수정한다.
"""
import sys, os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

CFG = json.load(open(sys.argv[1], encoding='utf-8'))
DEF = json.load(open(sys.argv[2], encoding='utf-8'))

WORK = CFG['work']['folder']
MAT  = os.path.join(WORK, CFG['work']['material_dir'])
OUT  = os.path.join(WORK, CFG['work']['sheet_name'])

if os.path.exists(OUT):
    print(f'중단: {OUT} 이 이미 있습니다.')
    print('덮어쓰면 사용자 입력이 사라집니다. read_workbook.py 로 읽어 부분 수정하세요.')
    sys.exit(1)

F = 'Arial'
IN_FILL   = PatternFill('solid', fgColor='FFF9E0')   # 사용자 입력칸
AUTO_FILL = PatternFill('solid', fgColor='E7F0FB')   # 몰 자동수집
HEAD_FILL = PatternFill('solid', fgColor='1F1F1F')
BOX = Border(*[Side('thin', color='D9D9D9')] * 4)
CEN = Alignment(horizontal='center', vertical='center')
WRAP = Alignment(vertical='center', wrap_text=True)

wb = Workbook()

# ---------------------------------------------------------------- ① 사용설명
ws = wb.active; ws.title = '① 사용설명'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 120
lines = [
    ('【 기획전 구성표 】 사용법', 16, True, '000000'),
    ('', 11, False, '000000'),
    ('1. [③ 섹션구성] 에서 노란 칸에 내용을 채워 주세요.', 11, True, '1F4E79'),
    ('   · 사용 = N 으로 바꾸면 그 구좌는 페이지에서 아예 사라집니다.', 10, False, '555555'),
    ('   · 내용을 비워두면 그 구좌도 그려지지 않습니다.', 10, False, '555555'),
    ('', 11, False, '000000'),
    ('2. 이미지가 필요한 구좌는 _소재 폴더에 파일을 넣어 주세요.', 11, True, '1F4E79'),
    ('   · 파일명은 "이미지 파일명" 칸과 같게 저장하시면 됩니다.', 10, False, '555555'),
    ('', 11, False, '000000'),
    ('3. 파란 칸은 몰에서 자동으로 가져온 값입니다. 손대지 않으셔도 됩니다.', 11, True, '1F4E79'),
    ('   · 상품 이미지·가격·할인율, 브랜드 로고·링크', 10, False, '555555'),
    ('', 11, False, '000000'),
    ('★ 저장하실 때는 엑셀을 반드시 닫아 주세요.', 12, True, 'C00000'),
    ('   열어둔 채로 두면 서로의 수정이 덮어써집니다.', 10, False, 'C00000'),
    ('', 11, False, '000000'),
    ('★ 채팅에 올라온 파일이 아니라, 이 폴더의 파일만 여세요.', 12, True, 'C00000'),
]
for i, (t, sz, bold, col) in enumerate(lines, 1):
    c = ws.cell(i, 2, t); c.font = Font(F, size=sz, bold=bold, color=col)

# ---------------------------------------------------------------- ② 공통설정
ws = wb.create_sheet('② 공통설정')
ws.column_dimensions['A'].width = 24; ws.column_dimensions['B'].width = 60
for j, h in enumerate(['항목', '값'], 1):
    c = ws.cell(1, j, h); c.font = Font(F, bold=True, color='FFFFFF', size=11)
    c.fill = HEAD_FILL; c.alignment = CEN; c.border = BOX
S = CFG['style']; P = CFG['promo']
common = [
    ('기획전명', P['title']), ('영문 표기', P.get('title_en', '')),
    ('진행 기간 시작', P['start']), ('진행 기간 종료', P['end']),
    ('에디터', P.get('editor', '')), ('최대 할인율', P.get('max_discount', '')),
    ('메인 컬러', S['main_color']), ('포인트 컬러', S['point_color']),
    ('쿠폰바 컬러', S['coupon_color']), ('섹션 교차 배경색', S['section_alt_bg']),
    ('본문 폰트', S['font_family']), ('컨텐츠 폭(px)', S['content_width']),
    ('상품 그리드 칼럼', S['grid_cols']), ('섹션당 기본 노출 상품수', S['default_show']),
    ('상품 링크 형식', CFG['mall']['product_view_url'].replace('{code}', '{상품코드}')),
]
for i, (k, v) in enumerate(common, 2):
    ws.cell(i, 1, k).font = Font(F, size=10, bold=True); ws.cell(i, 1).border = BOX
    c = ws.cell(i, 2, v); c.fill = IN_FILL; c.border = BOX; c.font = Font(F, size=10, color='0000FF')

# ---------------------------------------------------------------- ③ 섹션구성
ws = wb.create_sheet('③ 섹션구성')
HEAD = ['No', '사용', '섹션ID', '섹션명', '구좌ID', '구좌명', '항목',
        '유형', '내용', '이미지 파일명', '이미지 폴더', '권장 규격', '가이드']
WID  = [5, 6, 9, 16, 8, 16, 22, 9, 40, 20, 18, 20, 46]
for j, (h, w) in enumerate(zip(HEAD, WID), 1):
    c = ws.cell(1, j, h); c.font = Font(F, bold=True, color='FFFFFF', size=10)
    c.fill = HEAD_FILL; c.alignment = CEN; c.border = BOX
    ws.column_dimensions[chr(64 + j)].width = w

dv_use  = DataValidation(type='list', formula1='"Y,N"')
dv_type = DataValidation(type='list',
                         formula1='"텍스트,이미지,링크,색상,숫자,선택,상품코드,자동"')
ws.add_data_validation(dv_use); ws.add_data_validation(dv_type)

r, no = 2, 0
nav_sec = None
for sec in DEF['sections']:
    if sec.get('nav'):
        nav_sec = sec
    for slot in sec.get('slots', []):
        no += 1
        vals = [no, 'Y', sec['id'], sec['name'], slot['gid'], slot['gname'], slot['item'],
                slot.get('type', '텍스트'), slot.get('value', ''), slot.get('file', ''),
                sec.get('folder', ''), slot.get('spec', ''), slot.get('guide', '')]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v); c.border = BOX; c.font = Font(F, size=10)
            c.alignment = WRAP if j in (7, 13) else (CEN if j in (1, 2, 3, 5, 8) else None)
            if j in (2, 9, 10):
                c.fill = IN_FILL; c.font = Font(F, size=10, color='0000FF')
            if slot.get('type') == '자동' and j == 9:
                c.fill = AUTO_FILL
        r += 1

# 네비 항목 구좌 — 섹션 수만큼 Y/N 행 생성
if nav_sec:
    for sec in DEF['sections']:
        if sec.get('nav') or not sec.get('slots'):
            continue
        no += 1
        vals = [no, 'Y', nav_sec['id'], nav_sec['name'], 'B02', '네비 항목',
                f"{sec['name']} ({sec['id']})", '선택', 'Y', '', '', '',
                '이 메뉴만 네비게이션에서 빼려면 N. 섹션 내용은 그대로 남습니다.']
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v); c.border = BOX; c.font = Font(F, size=10)
            if j in (2, 9):
                c.fill = IN_FILL; c.font = Font(F, size=10, color='0000FF')
        r += 1

dv_use.add(f'B2:B{r+50}')
dv_type.add(f'H2:H{r+50}')

# ---------------------------------------------------------------- ④ 상품코드
ws = wb.create_sheet('④ 상품코드')
for j, (h, w) in enumerate(zip(['섹션ID', '섹션명', '사용', '노출 개수', '등록 수', '상품코드 (쉼표 구분)'],
                               [10, 20, 7, 11, 10, 120]), 1):
    c = ws.cell(1, j, h); c.font = Font(F, bold=True, color='FFFFFF', size=10)
    c.fill = HEAD_FILL; c.alignment = CEN; c.border = BOX
    ws.column_dimensions[chr(64 + j)].width = w
i = 2
for sec in DEF['sections']:
    if not sec.get('products'):
        continue
    ws.cell(i, 1, sec['id']).border = BOX
    ws.cell(i, 2, sec['name']).border = BOX
    for j, v in [(3, 'Y'), (4, S['default_show']), (5, 0), (6, '')]:
        c = ws.cell(i, j, v); c.fill = IN_FILL; c.border = BOX
        c.font = Font(F, size=10, color='0000FF')
    i += 1

# ---------------------------------------------------------------- ⑤ 참여브랜드
ws = wb.create_sheet('⑤ 참여브랜드')
for j, (h, w) in enumerate(zip(['No', '사용', '브랜드명', '몰 등록명 (자동)', '영문명 (자동)',
                                '최대 할인율', '브랜드ID (자동)', '로고 (자동)',
                                '브랜드관 링크 (자동)', '비고'],
                               [5, 6, 20, 20, 22, 12, 12, 44, 44, 26]), 1):
    c = ws.cell(1, j, h); c.font = Font(F, bold=True, color='FFFFFF', size=10)
    c.fill = HEAD_FILL; c.alignment = CEN; c.border = BOX
    ws.column_dimensions[chr(64 + j)].width = w
ws.cell(3, 1, '※ 브랜드명만 적어 주세요. 파란 칸은 몰에서 자동으로 채웁니다.').font = \
    Font(F, size=9, color='1F4E79')
ws.cell(4, 1, '※ 띄어쓰기가 달라도 몰에서 찾습니다. 표기가 다르면 주황색으로 표시됩니다.').font = \
    Font(F, size=9, color='1F4E79')

# ---------------------------------------------------------------- ⑥ 원본참고
ws = wb.create_sheet('⑥ 원본참고')
ws.column_dimensions['B'].width = 110
ws.cell(1, 2, '참고링크').font = Font(F, bold=True, size=12)
c = ws.cell(2, 2, DEF.get('reference_url', '')); c.fill = IN_FILL
ws.cell(4, 2, '분해 근거 / 메모').font = Font(F, bold=True, size=12)

# ---------------------------------------------------------------- ⑦ 화면 위치 안내
ws = wb.create_sheet('⑦ 화면 위치 안내 (크게)')
ws.sheet_view.showGridLines = False
ws.column_dimensions['B'].width = 150
ws.cell(1, 2, '화면 위치 안내 — 라벨 = [③ 섹션구성] 의 구좌ID · 항목').font = Font(F, bold=True, size=15)
ws.cell(3, 2, '_소재/00_위치안내/ 의 캡처를 여기에 넣습니다.').font = Font(F, size=10, color='888888')

wb.save(OUT)

# ---------------------------------------------------------------- 폴더 생성
os.makedirs(MAT, exist_ok=True)
made = []
for sec in DEF['sections']:
    f = sec.get('folder')
    if not f or f == '—':
        continue
    d = os.path.join(MAT, f)
    os.makedirs(d, exist_ok=True)
    made.append(f)
    g = os.path.join(d, '_규격안내.txt')
    if not os.path.exists(g):
        need = [s for s in sec.get('slots', []) if s.get('file')]
        body = f"[ {f} ] {sec['name']} 소재\n\n넣을 파일\n"
        body += ''.join(f"  {s['file']:20s} {s['item']:20s} {s.get('spec','')}\n" for s in need) \
                or "  (없음 — 이 폴더는 비워 두셔도 됩니다)\n"
        body += "\n메모\n  - 파일명은 엑셀 [③ 섹션구성] 의 \"이미지 파일명\" 칸과 같게 맞춰 주세요.\n"
        open(g, 'w', encoding='utf-8').write(body)

for extra in ['00_위치안내', '10_배너', '99_공통']:
    os.makedirs(os.path.join(MAT, extra), exist_ok=True)

print('생성:', OUT)
print('폴더:', ', '.join(made + ['00_위치안내', '10_배너', '99_공통']))
print('※ 여기서 멈춥니다. 사용자가 채우고 "엑셀 반영해줘" 라고 할 때까지 다음 단계로 가지 않습니다.')
